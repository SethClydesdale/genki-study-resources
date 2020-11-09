import re
import ast
import sys
from pathlib import Path
from itertools import chain

import openpyxl

lessons_folder = Path(sys.argv[1])
title_regex = re.compile(r'<title>(.*):(.*)- Lesson')
quizlet_regex = re.compile(r'quizlet : (.*?})', flags=re.S)
filter_regex = re.compile(r"format : 'kanji'|format : 'practice'|format : 'hirakata'|type : 'fill'|type : 'drawing'|type : 'stroke'|type : 'multi'|type : 'writing'|<title>Review:|Kanji Practice: Match the Readings|Kanji Practice: Match the Sentences|Kanji Practice: Match the Verbs|Katakana Practice: Countries and Capitals", flags=re.S)
output_folder = Path('.').absolute().joinpath('wordlist_E-J').joinpath(lessons_folder.name)


def get_tags(html):
    """
    <title>Useful Expressions: Time (Minutes 11-30) - Lesson 1 | Genki ...</title>
    Useful_Expressions , Time_(Minutes_11-30)
    """
    match = title_regex.search(html)
    return match.group(1).strip().replace(' ', '_'), match.group(2).strip().replace(' ', '_')


def get_vocab(html):
    return ast.literal_eval(quizlet_regex.search(html).group(1).replace(r'//', '#'))


def main():
    try :
        print('Creating folder for xlsx...')
        output_folder.mkdir(parents=True, exist_ok=False)
    except Exception:
        print('Folder already exists, skipping this step.')

    workbooks = list()

    for lesson_folder in lessons_folder.glob('lesson*'):
        lesson_number = lesson_folder.name.split('-')[-1]
        
        print(f'Getting vocab for Lesson {lesson_number}...')

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.cell(1, 1).value = "English"
        sheet.cell(1, 2).value = "Japanese"
        row_num = 2
        
        for vocab_folder in chain(lesson_folder.glob('vocab*'), lesson_folder.glob('literacy*')):
            with open(vocab_folder.joinpath('index.html'), 'r', encoding='UTF8') as f:
                html = f.read()
                if filter_regex.search(html) == None: # Filter out exercise types that are NOT vocab
                    try:
                        vocab = get_vocab(html)
                    except Exception:
                        print(f'Failed parsing of lesson-{lesson_number}, vocab file {vocab_folder}')
                        continue
                    for jp, eng in vocab.items():
                        eng = re.sub(r"\<(.*?)\>", '', eng)
                        sheet.cell(row_num, 1).value = eng
                        sheet.cell(row_num, 2).value = jp

                        row_num += 1

        workbooks.append((wb, lesson_folder.name))

    for wb, name in workbooks:
        print(f'Creating deck for {name}...');
        wb.save(output_folder.joinpath(f'{name}.xlsx'))
            
    print('All xlsx list for the selected edition have been generated!')


if __name__ == '__main__':
    main()